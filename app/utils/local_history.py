from __future__ import annotations

import json
import os
from datetime import datetime, UTC
from pathlib import Path
from typing import Dict, Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_HISTORY_PATH = Path("/app/data/issue_history.xlsx")
HISTORY_XLSX_PATH = Path(os.getenv("HISTORY_XLSX_PATH", DEFAULT_HISTORY_PATH))

HEADERS = [
    "Дата",
    # "Вкладка",
    "Ящик",
    "Айтем",
    "Количество",
    "Статус",
    "Ответственный",
    "Серийный номер",
    "Номер счёта",
]

def _format_datetime(dt: datetime) -> str:
    return f"{dt.day}.{dt.month}.{dt.year} | {dt.time().strftime('%H:%M:%S')}"

def _ensure_workbook():
    if HISTORY_XLSX_PATH.exists():
        try:
            return load_workbook(HISTORY_XLSX_PATH)
        except Exception:
            HISTORY_XLSX_PATH.unlink(missing_ok=True)
            
    wb = Workbook()
    ws = wb.active
    ws.title = "History"
    ws.append(HEADERS)
    for idx in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    return wb


def append_issue_row(data: Dict[str, Any]):
    """
    Добавляет запись об истории в локальный XLSX.
    data ожидает ключи:
      created_at (datetime), tab_name, box_name, item_name, qty, status, responsible, serial, invoice
    """
    # f"{now.day}.{now.month}.{now.year} | {now.time().strftime('%H:%M:%S')}"
    try:
        wb = _ensure_workbook()
        ws = wb.active
        created_at = data.get("created_at") or datetime.now(UTC)
        # formatted_created_at = f"{created_at.date()} | {created_at.time().strftime('%H:%M:%S')}"
        row = [
            _format_datetime(created_at),
            # data.get("tab_name") or "",
            data.get("box_name") or "",
            data.get("item_name") or "",
            data.get("qty") or "",
            data.get("status") or "",
            data.get("responsible") or "",
            data.get("serial") or "",
            data.get("invoice") or "",
        ]
        ws.append(row)
        HISTORY_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
        wb.save(HISTORY_XLSX_PATH)
    except Exception:
        # Логируем, но не блокируем основной поток
        import logging

        logging.getLogger(__name__).exception("Не удалось записать историю в XLSX")
